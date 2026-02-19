import json
import os
import queue
import sys
import threading
import winreg
import time
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox, ttk

import openpyxl


def _base_dir() -> str:
    """资源读取目录：打包后为 _MEIPASS，开发时为脚本目录"""
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))


def _work_dir() -> str:
    """可写工作目录：打包后为 exe 所在目录，开发时为脚本目录"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


# ---------------------------
# 账号管理
# ---------------------------
def load_accounts_from_excel(file_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    headers = [str(c.value).strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if 'username' not in headers or 'password' not in headers:
        raise ValueError("Excel 文件需包含 username 和 password 列")
    ui = headers.index('username')
    pi = headers.index('password')
    accounts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        u, p = row[ui], row[pi]
        if u and p:
            accounts.append({'username': str(u).strip(), 'password': str(p).strip()})
    return accounts


def save_accounts_to_config(accounts: list[dict]):
    config_path = os.path.join(_work_dir(), 'config.json')
    with open(config_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    data['accounts'] = accounts
    with open(config_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


def load_accounts_from_config() -> list[dict]:
    config_path = os.path.join(_work_dir(), 'config.json')
    with open(config_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return [a for a in data.get('accounts', []) if a.get('username')]


def _load_raw_config() -> dict:
    """加载原始 config.json（含注释字段）"""
    config_path = os.path.join(_work_dir(), 'config.json')
    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def _detect_steam_path() -> str:
    """自动检测 Steam 安装路径：优先注册表，其次遍历所有盘符常见目录"""
    # 1. 尝试注册表
    reg_keys = [
        (winreg.HKEY_CURRENT_USER, r"Software\Valve\Steam", "SteamPath"),
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Valve\Steam", "InstallPath"),
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\WOW6432Node\Valve\Steam", "InstallPath"),
    ]
    for hive, sub_key, value_name in reg_keys:
        try:
            with winreg.OpenKey(hive, sub_key) as key:
                val, _ = winreg.QueryValueEx(key, value_name)
                path = os.path.normpath(val)
                if os.path.isfile(os.path.join(path, "steam.exe")):
                    return path
        except OSError:
            continue

    # 2. 遍历所有盘符的常见安装路径
    import string
    drives = [f"{d}:\\" for d in string.ascii_uppercase if os.path.exists(f"{d}:\\")]
    sub_dirs = ["Steam", r"Program Files (x86)\Steam", r"Program Files\Steam"]
    for drive in drives:
        for sub in sub_dirs:
            p = os.path.join(drive, sub)
            if os.path.isfile(os.path.join(p, "steam.exe")):
                return p

    return ""


# ---------------------------
# 设置对话框
# ---------------------------
class SettingsDialog(tk.Toplevel):
    """配置设置对话框，分 Tab 展示所有 config.json 配置项"""

    def __init__(self, parent):
        super().__init__(parent)
        self.title("设置")
        self.geometry("520x480")
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)

        self._vars = {}  # 存储所有 StringVar / DoubleVar
        self._load_config()
        self._build_ui()

    def _load_config(self):
        """从 config.json 加载当前值"""
        raw = _load_raw_config()
        # 过滤注释
        def filt(d):
            if isinstance(d, dict):
                return {k: filt(v) for k, v in d.items() if not k.startswith('__comment__')}
            if isinstance(d, list):
                return [filt(i) for i in d]
            return d
        self._cfg = filt(raw)

    def _build_ui(self):
        nb = ttk.Notebook(self)
        nb.pack(fill=tk.BOTH, expand=True, padx=8, pady=(8, 0))

        nb.add(self._build_steam_tab(nb), text="Steam")
        nb.add(self._build_paths_tab(nb), text="路径")
        nb.add(self._build_timing_tab(nb), text="时间")
        nb.add(self._build_buttons_tab(nb), text="按钮")

        # 底部按钮
        btn_frame = tk.Frame(self)
        btn_frame.pack(fill=tk.X, padx=8, pady=8)
        tk.Button(btn_frame, text="保存", width=10, command=self._save).pack(side=tk.RIGHT, padx=(4, 0))
        tk.Button(btn_frame, text="取消", width=10, command=self.destroy).pack(side=tk.RIGHT)

    # ---------- 辅助方法 ----------
    def _add_entry(self, parent, row, label, key, value, width=38):
        """添加一行 Label + Entry，返回 StringVar"""
        tk.Label(parent, text=label, anchor="w").grid(row=row, column=0, sticky="w", padx=6, pady=3)
        var = tk.StringVar(value=str(value))
        tk.Entry(parent, textvariable=var, width=width).grid(row=row, column=1, sticky="w", padx=6, pady=3)
        self._vars[key] = var
        return var

    def _add_range_entry(self, parent, row, label, key, values, width=12):
        """添加一行 Label + 两个 Entry（最小/最大），用于范围配置"""
        tk.Label(parent, text=label, anchor="w").grid(row=row, column=0, sticky="w", padx=6, pady=3)
        frame = tk.Frame(parent)
        frame.grid(row=row, column=1, sticky="w", padx=6, pady=3)
        var_min = tk.StringVar(value=str(values[0]))
        var_max = tk.StringVar(value=str(values[1]))
        tk.Entry(frame, textvariable=var_min, width=width).pack(side=tk.LEFT)
        tk.Label(frame, text=" ~ ").pack(side=tk.LEFT)
        tk.Entry(frame, textvariable=var_max, width=width).pack(side=tk.LEFT)
        self._vars[key + "_min"] = var_min
        self._vars[key + "_max"] = var_max

    # ---------- Tab 构建 ----------
    def _build_steam_tab(self, parent):
        frame = tk.Frame(parent)
        steam = self._cfg.get("steam", {})

        # Steam 路径 + 自动检测按钮
        tk.Label(frame, text="Steam 路径", anchor="w").grid(row=0, column=0, sticky="w", padx=6, pady=3)
        path_var = tk.StringVar(value=str(steam.get("path", "")))
        tk.Entry(frame, textvariable=path_var, width=30).grid(row=0, column=1, sticky="w", padx=6, pady=3)
        self._vars["steam.path"] = path_var

        def _auto_detect():
            found = _detect_steam_path()
            if found:
                exe = os.path.join(found, "steam.exe")
                path_var.set(exe if os.path.isfile(exe) else found)
            else:
                from tkinter import messagebox
                messagebox.showwarning("提示", "未能自动检测到 Steam 安装路径", parent=self)

        tk.Button(frame, text="自动检测", command=_auto_detect).grid(row=0, column=2, padx=4, pady=3)

        self._add_entry(frame, 1, "游戏 App ID", "steam.game_id", steam.get("game_id", ""))
        self._add_entry(frame, 2, "游戏进程名", "steam.process_name", steam.get("process_name", ""))
        self._add_entry(frame, 3, "子进程列表（逗号分隔）", "steam.child_processes",
                        ", ".join(steam.get("child_processes", [])))
        return frame

    def _build_paths_tab(self, parent):
        frame = tk.Frame(parent)
        paths = self._cfg.get("paths", {})
        self._add_entry(frame, 0, "封禁截图目录", "paths.screenshots", paths.get("screenshots", ""))
        self._add_entry(frame, 1, "调试截图目录", "paths.debug", paths.get("debug", ""))
        self._add_entry(frame, 2, "按钮截图目录", "paths.button_images", paths.get("button_images", ""))
        return frame

    def _build_timing_tab(self, parent):
        frame = tk.Frame(parent)
        timing = self._cfg.get("timing", {})
        r = 0
        self._add_entry(frame, r, "登录等待时间(秒)", "timing.login_delay",
                        timing.get("login_delay", 15), width=12); r += 1

        pc = timing.get("process_check", {})
        self._add_entry(frame, r, "进程启动超时(秒)", "timing.process_check.timeout",
                        pc.get("timeout", 120), width=12); r += 1
        self._add_entry(frame, r, "进程检查间隔(秒)", "timing.process_check.interval",
                        pc.get("interval", 5), width=12); r += 1

        ag = timing.get("agreement", {})
        self._add_entry(frame, r, "协议搜索耗时(秒)", "timing.agreement.search_time",
                        ag.get("search_time", 3), width=12); r += 1
        self._add_entry(frame, r, "图像匹配置信度", "timing.agreement.confidence",
                        ag.get("confidence", 0.85), width=12); r += 1

        delays = timing.get("delays", {})
        self._add_range_entry(frame, r, "随机延迟(秒)", "timing.delays.random",
                              delays.get("random", [2, 5])); r += 1
        self._add_range_entry(frame, r, "点击后延迟(秒)", "timing.delays.post_click",
                              delays.get("post_click", [0.5, 1.2])); r += 1
        self._add_range_entry(frame, r, "游戏加载等待(秒)", "timing.delays.game_loading",
                              delays.get("game_loading", [10, 15])); r += 1

        fv = timing.get("final_validation", {})
        self._add_range_entry(frame, r, "最终验证等待(秒)", "timing.final_validation.delay_range",
                              fv.get("delay_range", [20, 25])); r += 1
        return frame

    def _build_buttons_tab(self, parent):
        frame = tk.Frame(parent)
        buttons = self._cfg.get("buttons", {})
        r = 0
        self._add_entry(frame, r, "协议同意按钮图片", "buttons.eula_agree",
                        buttons.get("eula_agree", "")); r += 1
        self._add_entry(frame, r, "开始游戏按钮图片", "buttons.start_game",
                        buttons.get("start_game", "")); r += 1
        self._add_entry(frame, r, "二次协议按钮图片", "buttons.agreement",
                        buttons.get("agreement", "")); r += 1

        # start_game 按钮配置
        tk.Label(frame, text="── 开始游戏按钮配置 ──", font=("", 9, "bold")).grid(
            row=r, column=0, columnspan=2, sticky="w", padx=6, pady=(10, 3)); r += 1
        sg_cfg = buttons.get("config", {}).get("start_game", {})
        self._add_entry(frame, r, "超时(秒)", "buttons.config.start_game.timeout",
                        sg_cfg.get("timeout", 60), width=12); r += 1
        self._add_entry(frame, r, "重试次数", "buttons.config.start_game.retries",
                        sg_cfg.get("retries", 3), width=12); r += 1
        self._add_range_entry(frame, r, "点击后延迟(秒)", "buttons.config.start_game.post_delay",
                              sg_cfg.get("post_delay", [2, 3])); r += 1

        # agreement 按钮配置
        tk.Label(frame, text="── 协议确认按钮配置 ──", font=("", 9, "bold")).grid(
            row=r, column=0, columnspan=2, sticky="w", padx=6, pady=(10, 3)); r += 1
        ag_cfg = buttons.get("config", {}).get("agreement", {})
        self._add_entry(frame, r, "超时(秒)", "buttons.config.agreement.timeout",
                        ag_cfg.get("timeout", 20), width=12); r += 1
        self._add_entry(frame, r, "重试次数", "buttons.config.agreement.retries",
                        ag_cfg.get("retries", 2), width=12); r += 1
        self._add_range_entry(frame, r, "点击后延迟(秒)", "buttons.config.agreement.post_delay",
                              ag_cfg.get("post_delay", [2, 3])); r += 1

        # optional 复选框
        self._agreement_optional = tk.BooleanVar(value=ag_cfg.get("optional", True))
        tk.Checkbutton(frame, text="协议步骤可选（跳过不报错）",
                       variable=self._agreement_optional).grid(
            row=r, column=0, columnspan=2, sticky="w", padx=6, pady=3)
        return frame

    # ---------- 保存 ----------
    def _save(self):
        """收集所有字段值，写入 config.json 并重载"""
        try:
            v = self._vars

            def _float(key):
                return float(v[key].get())

            def _int(key):
                return int(float(v[key].get()))

            def _str(key):
                return v[key].get().strip()

            def _range(key):
                return [float(v[key + "_min"].get()), float(v[key + "_max"].get())]

            updates = {
                "steam": {
                    "path": _str("steam.path"),
                    "game_id": _str("steam.game_id"),
                    "process_name": _str("steam.process_name"),
                    "child_processes": [s.strip() for s in _str("steam.child_processes").split(",") if s.strip()],
                },
                "paths": {
                    "screenshots": _str("paths.screenshots"),
                    "debug": _str("paths.debug"),
                    "button_images": _str("paths.button_images"),
                },
                "timing": {
                    "login_delay": _float("timing.login_delay"),
                    "process_check": {
                        "timeout": _float("timing.process_check.timeout"),
                        "interval": _float("timing.process_check.interval"),
                    },
                    "agreement": {
                        "search_time": _float("timing.agreement.search_time"),
                        "confidence": _float("timing.agreement.confidence"),
                    },
                    "delays": {
                        "random": _range("timing.delays.random"),
                        "post_click": _range("timing.delays.post_click"),
                        "game_loading": _range("timing.delays.game_loading"),
                    },
                    "final_validation": {
                        "delay_range": _range("timing.final_validation.delay_range"),
                    },
                },
                "buttons": {
                    "eula_agree": _str("buttons.eula_agree"),
                    "start_game": _str("buttons.start_game"),
                    "agreement": _str("buttons.agreement"),
                    "config": {
                        "start_game": {
                            "timeout": _int("buttons.config.start_game.timeout"),
                            "retries": _int("buttons.config.start_game.retries"),
                            "post_delay": _range("buttons.config.start_game.post_delay"),
                        },
                        "agreement": {
                            "timeout": _int("buttons.config.agreement.timeout"),
                            "retries": _int("buttons.config.agreement.retries"),
                            "post_delay": _range("buttons.config.agreement.post_delay"),
                            "optional": self._agreement_optional.get(),
                        },
                    },
                },
            }

            from config_loader import ConfigLoader
            ConfigLoader.save_full_config(updates)
            messagebox.showinfo("提示", "设置已保存", parent=self)
            self.destroy()

        except ValueError as e:
            messagebox.showerror("输入错误", f"请检查数值格式:\n{e}", parent=self)
        except Exception as e:
            messagebox.showerror("保存失败", str(e), parent=self)


# ---------------------------
# 主界面
# ---------------------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Delta Steam Verifier")
        self.geometry("820x560")
        self.resizable(False, False)
        self._log_queue: queue.Queue = queue.Queue()
        self._running = False
        self._build_ui()
        self._load_accounts()
        self._poll_log()

    def _build_ui(self):
        # 左侧账号面板
        left = tk.Frame(self, width=320)
        left.pack(side=tk.LEFT, fill=tk.Y, padx=(10, 5), pady=10)
        left.pack_propagate(False)

        tk.Label(left, text="账号列表", font=("", 11, "bold")).pack(anchor="w")

        # 账号表格（包在子 frame 中，避免 side=LEFT 抢占按钮行空间）
        tree_frame = tk.Frame(left)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        cols = ("username", "password")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=18)
        self.tree.heading("username", text="账号")
        self.tree.heading("password", text="密码")
        self.tree.column("username", width=140)
        self.tree.column("password", width=140)
        sb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        # 按钮行
        btn_frame = tk.Frame(left)
        btn_frame.pack(fill=tk.X, pady=(6, 0))

        # 右侧操作面板
        right = tk.Frame(self)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 10), pady=10)

        tk.Button(btn_frame, text="导入 Excel", width=10, command=self._import_excel).pack(side=tk.LEFT, padx=(0, 4))
        tk.Button(btn_frame, text="手动添加", width=10, command=self._add_account).pack(side=tk.LEFT, padx=(0, 4))
        tk.Button(btn_frame, text="删除选中", width=10, command=self._delete_selected).pack(side=tk.LEFT)

        # 日志区
        tk.Label(right, text="运行日志", font=("", 11, "bold")).pack(anchor="w")
        self.log_box = tk.Text(right, state=tk.DISABLED, bg="#1e1e1e", fg="#d4d4d4",
                               font=("Courier", 10), wrap=tk.WORD)
        self.log_box.pack(fill=tk.BOTH, expand=True)

        log_sb = ttk.Scrollbar(right, orient=tk.VERTICAL, command=self.log_box.yview)
        self.log_box.configure(yscrollcommand=log_sb.set)
        log_sb.place(relx=1, rely=0, relheight=1, anchor="ne")

        # 底部控制栏
        ctrl = tk.Frame(right)
        ctrl.pack(fill=tk.X, pady=(6, 0))

        self.start_btn = tk.Button(ctrl, text="▶  开始运行", bg="#0e7a0d", fg="white",
                                   font=("", 10, "bold"), width=14, command=self._start)
        self.start_btn.pack(side=tk.LEFT)

        self.stop_btn = tk.Button(ctrl, text="■  停止", bg="#a00", fg="white",
                                  font=("", 10, "bold"), width=10,
                                  state=tk.DISABLED, command=self._stop)
        self.stop_btn.pack(side=tk.LEFT, padx=(8, 0))

        tk.Button(ctrl, text="清空日志", command=self._clear_log).pack(side=tk.RIGHT)
        tk.Button(ctrl, text="⚙ 设置", command=self._open_settings).pack(side=tk.RIGHT, padx=(0, 8))

    # ---------------------------
    # 账号操作
    # ---------------------------
    def _load_accounts(self):
        try:
            accounts = load_accounts_from_config()
            self.tree.delete(*self.tree.get_children())
            for a in accounts:
                self.tree.insert("", tk.END, values=(a['username'], a['password']))
        except Exception as e:
            messagebox.showerror("错误", f"加载账号失败: {e}")

    def _import_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx *.xls")])
        if not path:
            return
        try:
            accounts = load_accounts_from_excel(path)
            for a in accounts:
                self.tree.insert("", tk.END, values=(a['username'], a['password']))
            self._sync_accounts()
            self._log(f"已导入 {len(accounts)} 个账号")
        except Exception as e:
            messagebox.showerror("导入失败", str(e))

    def _add_account(self):
        win = tk.Toplevel(self)
        win.title("添加账号")
        win.geometry("280x130")
        win.resizable(False, False)
        win.grab_set()

        tk.Label(win, text="账号").grid(row=0, column=0, padx=10, pady=8, sticky="e")
        u_var = tk.StringVar()
        tk.Entry(win, textvariable=u_var, width=22).grid(row=0, column=1, padx=10)

        tk.Label(win, text="密码").grid(row=1, column=0, padx=10, pady=4, sticky="e")
        p_var = tk.StringVar()
        tk.Entry(win, textvariable=p_var, show="*", width=22).grid(row=1, column=1, padx=10)

        def confirm():
            u, p = u_var.get().strip(), p_var.get().strip()
            if not u or not p:
                messagebox.showwarning("提示", "账号和密码不能为空", parent=win)
                return
            self.tree.insert("", tk.END, values=(u, p))
            self._sync_accounts()
            win.destroy()

        tk.Button(win, text="确认", command=confirm, width=10).grid(row=2, column=0, columnspan=2, pady=10)

    def _delete_selected(self):
        selected = self.tree.selection()
        if not selected:
            return
        for item in selected:
            self.tree.delete(item)
        self._sync_accounts()

    def _sync_accounts(self):
        accounts = [
            {'username': str(self.tree.item(i)['values'][0]),
             'password': str(self.tree.item(i)['values'][1])}
            for i in self.tree.get_children()
        ]
        save_accounts_to_config(accounts)

    # ---------------------------
    # 运行控制
    # ---------------------------
    def _start(self):
        accounts = [
            {'username': str(self.tree.item(i)['values'][0]),
             'password': str(self.tree.item(i)['values'][1])}
            for i in self.tree.get_children()
        ]
        if not accounts:
            messagebox.showwarning("提示", "请先添加账号")
            return

        self._running = True
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self._log("=== 开始运行 ===")

        self._thread = threading.Thread(target=self._run_task, args=(accounts,), daemon=True)
        self._thread.start()

    def _stop(self):
        self._running = False
        self._log("=== 用户已停止 ===")
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)

    def _run_task(self, accounts):
        try:
            import main as core
            from config_loader import ConfigLoader
            ConfigLoader.reload()  # 确保读取最新配置
            core._log_callback = self._log
            core._should_stop = lambda: not self._running  # 注入停止检查函数

            for idx, acc in enumerate(accounts):
                if not self._running:
                    break
                self._log(f"\n{'='*36}")
                self._log(f"处理账号 ({idx+1}/{len(accounts)}): {acc['username']}")
                try:
                    core.login_steam(acc['username'], acc['password'])
                    core.launch_game()
                    time.sleep(15)
                    ban_info = core.capture_ban_info(acc['username'])

                    log_msg = f"{datetime.now()}, {acc['username']}, "
                    if ban_info["is_banned"]:
                        log_msg += f"封禁 | 时长:{ban_info['duration']} | 解封:{ban_info['unban_time']}"
                        self._log(f"!!! 封禁警报 !!! {ban_info['raw_text'][:50]}")
                    else:
                        log_msg += "正常"
                        self._log("账号状态：正常")

                    with open("result.log", "a", encoding="utf-8") as f:
                        f.write(log_msg + "\n")
                except Exception as e:
                    self._log(f"流程异常: {e}")
                finally:
                    core.logout_steam()
        finally:
            self._running = False
            self.after(0, lambda: self.start_btn.config(state=tk.NORMAL))
            self.after(0, lambda: self.stop_btn.config(state=tk.DISABLED))
            self._log("=== 运行结束 ===")

    # ---------------------------
    # 日志
    # ---------------------------
    def _log(self, msg: str):
        self._log_queue.put(msg)

    def _poll_log(self):
        while not self._log_queue.empty():
            msg = self._log_queue.get_nowait()
            self.log_box.config(state=tk.NORMAL)
            self.log_box.insert(tk.END, msg + "\n")
            self.log_box.see(tk.END)
            self.log_box.config(state=tk.DISABLED)
        self.after(100, self._poll_log)

    def _clear_log(self):
        self.log_box.config(state=tk.NORMAL)
        self.log_box.delete("1.0", tk.END)
        self.log_box.config(state=tk.DISABLED)

    def _open_settings(self):
        SettingsDialog(self)


if __name__ == "__main__":
    import shutil
    import pytesseract

    # 打包后首次运行：将内嵌资源复制到 exe 同级目录
    if getattr(sys, 'frozen', False):
        base = _base_dir()   # _MEIPASS（内嵌资源）
        work = _work_dir()   # exe 所在目录（可写）

        # config.json：仅在不存在时复制（保留用户修改）
        dst_cfg = os.path.join(work, 'config.json')
        if not os.path.exists(dst_cfg):
            shutil.copy2(os.path.join(base, 'config.json'), dst_cfg)

        # images 目录
        dst_images = os.path.join(work, 'images')
        if not os.path.exists(dst_images):
            shutil.copytree(os.path.join(base, 'images'), dst_images)

    # 设置 Tesseract 路径
    pytesseract.pytesseract.tesseract_cmd = os.path.join(
        _base_dir(), "Tesseract-OCR", "tesseract.exe"
    )

    # 捕获配置加载失败，弹窗提示而不是静默崩溃
    try:
        from config_loader import config  # noqa: F401 触发初始化
    except Exception as e:
        import tkinter as _tk
        _r = _tk.Tk()
        _r.withdraw()
        from tkinter import messagebox as _mb
        _mb.showerror("配置错误", f"config.json 加载失败:\n{e}")
        _r.destroy()
        raise SystemExit(1)

    app = App()
    app.mainloop()
